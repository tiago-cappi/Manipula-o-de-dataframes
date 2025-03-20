# No início do arquivo, após as importações existentes
import pandas as pd
import streamlit as st
import numpy as np
import io
import datetime
import re
import gc
from datetime import datetime

# Importe para processar os dados conforme o arquivo análise_produtos_clientes.py
from analise_pendentes import exibir_analise_pendentes


# Adicione esta função para replicar a lógica do análise_produtos_clientes.py
def processar_dados_produtos_clientes(df_analise, df_categorias):
    """
    Processa os dados de análise comercial e categoria de produtos para gerar análise
    de produtos por cliente com histórico de interações conforme arquivo análise_produtos_clientes.py.
    
    Args:
        df_analise: DataFrame com dados de análise comercial
        df_categorias: DataFrame com dados de categorias de produtos
        
    Returns:
        DataFrame processado com a análise por cliente/produto
    """
    try:
        st.write("Iniciando processamento de dados conforme análise_produtos_clientes.py...")
        
        # 1. Classificar clientes ABC
        df_classificacao_abc = classificar_clientes_abc(df_analise)
        
        # 2. Realizar merge entre df_analise e df_classificacao_abc
        df_resultado = pd.merge(
            df_analise[["Código Produto", "Descrição Produto", "Dt Entrada", "Cliente", 
                        "Consultor Interno", "Prob.Fech.", "Motivo Não Venda"]],
            df_classificacao_abc,
            on='Cliente', 
            how='inner'
        )
        
        # 3. Juntar com as categorias de produtos
        df_categorias_slim = df_categorias[["Código Produto", "Negócio", "Grupo", "Subgrupo"]]
        df_resultado_final = pd.merge(df_resultado, df_categorias_slim, on="Código Produto", how="left")
        
        # 4. Converter coluna 'Dt Entrada' para datetime
        df_resultado_final['Dt Entrada_temp'] = pd.to_datetime(df_resultado_final['Dt Entrada'], errors='coerce')
        
        # 5. Agrupar por subgrupo, código produto e cliente
        resultado = []
        
        for (subgrupo, codigo_produto, cliente), grupo in df_resultado_final.groupby(["Subgrupo", "Código Produto", "Cliente"]):
            # Ordenar dados por data
            grupo_ordenado = grupo.sort_values("Dt Entrada_temp")
            
            # Criar linha para o resultado
            linha = {
                "Subgrupo": subgrupo,
                "Negócio": grupo_ordenado["Negócio"].iloc[0] if "Negócio" in grupo_ordenado.columns else "",
                "Grupo": grupo_ordenado["Grupo"].iloc[0] if "Grupo" in grupo_ordenado.columns else "",
                "Código Produto": codigo_produto,
                "Descrição Produto": grupo_ordenado["Descrição Produto"].iloc[0],
                "Cliente": cliente,
                "Nome Cliente": grupo_ordenado["Nome Cliente"].iloc[0],
                "UF": grupo_ordenado["UF"].iloc[0],
                "Cidade": grupo_ordenado["Cidade"].iloc[0],
                "ABC": grupo_ordenado["ABC"].iloc[0],
                "Valor Total Orçado": grupo_ordenado["Valor Total Orçado"].iloc[0]
            }
            
            # Adicionar histórico de interações
            linha["Dt Entrada"] = grupo_ordenado["Dt Entrada_temp"].dt.strftime("%Y-%m-%d").tolist()
            linha["Prob.Fech."] = grupo_ordenado["Prob.Fech."].tolist()
            linha["Motivo Não Venda"] = grupo_ordenado["Motivo Não Venda"].tolist()
            
            # Calcular última data e consultor
            if len(grupo_ordenado) > 0:
                ultima_data_idx = grupo_ordenado["Dt Entrada_temp"].idxmax()
                linha["Última Data"] = grupo_ordenado.loc[ultima_data_idx, "Dt Entrada_temp"]
                linha["Último Consultor"] = grupo_ordenado.loc[ultima_data_idx, "Consultor Interno"]
            
            resultado.append(linha)
        
        # 6. Criar DataFrame final
        df_final = pd.DataFrame(resultado)
        
        # 7. Formatar datas
        if "Última Data" in df_final.columns:
            df_final["Última Data"] = pd.to_datetime(df_final["Última Data"]).dt.strftime("%Y-%m-%d")
        
        st.success(f"Processamento concluído! {len(df_final)} registros gerados.")
        return df_final
        
    except Exception as e:
        st.error(f"Erro ao processar dados de produtos por cliente: {e}")
        import traceback
        st.error(traceback.format_exc())
        return pd.DataFrame()




if 'tab_loaded' not in st.session_state:
    st.session_state.tab_loaded = {
        'tab1': False,
        'tab2': False,
        'tab3': False,
        'tab4': False
    }


# Set page title and layout
st.set_page_config(page_title="Dashboard de Dados Comerciais", layout="wide")

# Title
st.title("Dashboard de Análise Comercial")




# Adicione esta função antes de carregar_dados
@st.cache_data
def carregar_excel_corretamente(arquivo, header_row=0):
    """
    Carrega um arquivo Excel garantindo que os cabeçalhos sejam interpretados corretamente.
    
    Args:
        arquivo: Caminho do arquivo ou objeto de arquivo
        header_row: Índice da linha que contém os nomes das colunas (0 por padrão)
    
    Returns:
        DataFrame do pandas carregado corretamente
    """
    try:
        # Mostrar mensagem de carregamento
        with st.spinner('Carregando arquivo Excel...'):
            # Verificar se é um objeto de arquivo ou caminho
            if hasattr(arquivo, 'read'):
                # É um objeto de arquivo (upload via Streamlit)
                # Primeiro, verificamos os nomes das colunas
                df_header = pd.read_excel(arquivo, nrows=0)
                st.write(f"Colunas detectadas: {list(df_header.columns)}")
                
                # Reiniciar o ponteiro do arquivo
                arquivo.seek(0)
                
                # Carregar o arquivo completo com os cabeçalhos corretos
                df = pd.read_excel(
                    arquivo,
                    header=header_row
                )
            else:
                # É um caminho de arquivo
                df = pd.read_excel(
                    arquivo,
                    header=header_row
                )
                
            # Mostrar informações sobre o DataFrame carregado
            st.success(f"Arquivo carregado com sucesso: {len(df)} linhas, {len(df.columns)} colunas")
            
            # Verificar se o número de colunas está razoável
            if len(df.columns) > 100:
                st.warning(f"Detectado um número anormalmente alto de colunas: {len(df.columns)}. Verificando possível erro de formatação...")
                
                # Tentar identificar o problema e corrigir
                if isinstance(header_row, int) and header_row == 0:
                    # Tentar carregar com diferentes opções de cabeçalho
                    if hasattr(arquivo, 'read'):
                        arquivo.seek(0)
                    
                    # Mostrar as primeiras linhas para análise
                    df_preview = pd.read_excel(arquivo, nrows=5, header=None)
                    st.write("Visualização das primeiras linhas (sem cabeçalho):")
                    st.dataframe(df_preview)
                    
                    # Sugerir correção
                    st.warning("Parece que há um problema com o cabeçalho do arquivo. Você pode:")
                    st.info("1. Verifique se o arquivo está no formato correto (tabular)")
                    st.info("2. Tente carregar novamente especificando qual linha contém os cabeçalhos (0-indexado)")
                    
                    # Oferecer interface para o usuário escolher a linha de cabeçalho
                    novo_header = st.number_input("Linha do cabeçalho (0 é a primeira linha):", 0, 10, 0)
                    
                    if novo_header != header_row and st.button("Recarregar com novo cabeçalho"):
                        # Recarregar com o novo header
                        if hasattr(arquivo, 'read'):
                            arquivo.seek(0)
                        return carregar_excel_corretamente(arquivo, header_row=novo_header)
            
            return df
                
    except Exception as e:
        st.error(f"Erro ao carregar arquivo Excel: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None




# Cache the data loading function to improve performance
@st.cache_data
def carregar_dados(caminho_arquivo):
    """
    Lê um arquivo Excel e o carrega como um DataFrame do Pandas.
    """
    try:
        df = pd.read_excel(caminho_arquivo)
        return df
    except Exception as e:
        st.error(f"Erro ao ler o arquivo Excel: {e}")
        return None
    
    
# Adicione esta função após a função carregar_dados
@st.cache_data(ttl=3600)
def otimizar_dataframe_inicial(df):
    """Reduzir o dataframe inicial para melhorar performance"""
    # Converter tipos de dados para otimizar memória
    for col in df.select_dtypes(include=['object']).columns:
        if df[col].nunique() < df.shape[0] / 2:  # Se coluna tem cardinalidade baixa
            df[col] = df[col].astype('category')
    
    # Converter colunas numéricas para tipos mais eficientes
    for col in df.select_dtypes(include=['float']).columns:
        df[col] = pd.to_numeric(df[col], downcast='float')
    
    for col in df.select_dtypes(include=['int']).columns:
        df[col] = pd.to_numeric(df[col], downcast='integer')
    
    return df



# Function to process data
# Substitua a função processar_dados existente pela nova abaixo
@st.cache_data(ttl=3600, max_entries=10)
def processar_dados(df_analise, df_categorias):
    """Processa os dados comerciais para análise."""
    try:
        # Verificar se temos dados suficientes
        if df_analise is None or df_categorias is None or len(df_analise) == 0 or len(df_categorias) == 0:
            st.error("Dados insuficientes para processamento.")
            return pd.DataFrame()
            
        # 1. Mostrar informações de diagnóstico
        st.write(f"DataFrame de análise: {df_analise.shape[0]} linhas x {df_analise.shape[1]} colunas")
        st.write(f"DataFrame de categorias: {df_categorias.shape[0]} linhas x {df_categorias.shape[1]} colunas")
        
        # 2. Verificar colunas necessárias no DataFrame de análise
        colunas_essenciais_analise = ["Cliente", "Código Produto", "Dt Entrada", "Valor Orçado", 
                                      "Nome Cliente", "Consultor Interno"]
        
        colunas_faltantes = [col for col in colunas_essenciais_analise if col not in df_analise.columns]
        if colunas_faltantes:
            st.warning(f"Colunas essenciais faltando no DataFrame de análise: {', '.join(colunas_faltantes)}")
            st.warning("Processamento pode não funcionar corretamente sem estas colunas.")
        
        # 3. Verificar colunas necessárias no DataFrame de categorias
        colunas_categorias = ["Código Produto", "Negócio", "Grupo", "Subgrupo"]
        colunas_faltantes_cat = [col for col in colunas_categorias if col not in df_categorias.columns]
        if colunas_faltantes_cat:
            st.warning(f"Colunas essenciais faltando no DataFrame de categorias: {', '.join(colunas_faltantes_cat)}")
            st.warning("Informações de categorização podem estar incompletas.")
        
        # 4. Converter a coluna de data para datetime
        if 'Dt Entrada' in df_analise.columns:
            df_analise['Dt Entrada'] = pd.to_datetime(df_analise['Dt Entrada'], errors='coerce')
        
        # 5. Classificar clientes ABC
        st.write("Iniciando classificação ABC de clientes...")
        df_clientes_abc = classificar_clientes_abc(df_analise)
        
        # 6. Criar dicionário de categorias de produtos
        st.write("Criando dicionário de categorias de produtos...")
        categorias_dict = {}
        
        # Adaptar para as colunas disponíveis no DataFrame de categorias
        colunas_categoria_disponiveis = ["Negócio", "Grupo", "Subgrupo"]
        colunas_disponiveis = [col for col in colunas_categoria_disponiveis if col in df_categorias.columns]
        
        for _, row in df_categorias.iterrows():
            if "Código Produto" in df_categorias.columns:
                cod_produto = str(row["Código Produto"])
                categorias_dict[cod_produto] = {}
                
                for col in colunas_disponiveis:
                    categorias_dict[cod_produto][col] = row.get(col, "")
        
        # 7. Criar DataFrame final com informações combinadas
        st.write("Combinando informações de produtos e clientes...")
        
        resultado = []
        total_registros = len(df_analise)
        
        # Criar barra de progresso
        progress_bar = st.progress(0)
        status_texto = st.empty()
        
        # Processar em lotes para economizar memória
        lote_size = 1000
        for i in range(0, total_registros, lote_size):
            # Atualizar progresso
            progress = min(i / total_registros, 1.0)
            progress_bar.progress(progress)
            status_texto.text(f"Processando registros {i+1}-{min(i+lote_size, total_registros)} de {total_registros}")
            
            # Processar lote atual
            lote = df_analise.iloc[i:i+lote_size]
            
            # Agrupar por produto e cliente
            for (cliente, codigo_produto), grupo in lote.groupby(["Cliente", "Código Produto"]):
                # Informações do produto
                produto_info = {}
                if "Descrição Produto" in grupo.columns:
                    produto_info["Descrição Produto"] = grupo["Descrição Produto"].iloc[0]
                else:
                    produto_info["Descrição Produto"] = ""
                
                # Verificar se o produto existe no dicionário de categorias
                if codigo_produto in categorias_dict:
                    for cat_col in colunas_disponiveis:
                        produto_info[cat_col] = categorias_dict[codigo_produto][cat_col]
                else:
                    for cat_col in colunas_disponiveis:
                        produto_info[cat_col] = ""
                
                # Informações do cliente
                cliente_info = {}
                if "Nome Cliente" in grupo.columns:
                    cliente_info["Nome Cliente"] = grupo["Nome Cliente"].iloc[0]
                else:
                    cliente_info["Nome Cliente"] = ""
                
                # Verificar se o cliente existe no DataFrame ABC
                cliente_row = df_clientes_abc[df_clientes_abc["Cliente"] == cliente] if len(df_clientes_abc) > 0 else None
                if cliente_row is not None and len(cliente_row) > 0:
                    cliente_info["ABC"] = cliente_row["ABC"].iloc[0]
                    cliente_info["UF"] = cliente_row["UF"].iloc[0] if "UF" in cliente_row.columns else ""
                    cliente_info["Cidade"] = cliente_row["Cidade"].iloc[0] if "Cidade" in cliente_row.columns else ""
                    cliente_info["Valor Total Orçado"] = cliente_row["Valor Total Orçado"].iloc[0]
                else:
                    cliente_info["ABC"] = "C"
                    cliente_info["UF"] = ""
                    cliente_info["Cidade"] = ""
                    cliente_info["Valor Total Orçado"] = 0
                
                # Histórico de interações
                if "Dt Entrada" in grupo.columns:
                    historico_datas = grupo["Dt Entrada"].dt.strftime("%Y-%m-%d").tolist()
                else:
                    historico_datas = []
                
                historico_prob = grupo["Prob.Fech."].tolist() if "Prob.Fech." in grupo.columns else []
                historico_motivo = grupo["Motivo Não Venda"].tolist() if "Motivo Não Venda" in grupo.columns else []
                
                # Última interação
                if "Dt Entrada" in grupo.columns and len(grupo) > 0:
                    ultima_idx = grupo["Dt Entrada"].idxmax() if not grupo["Dt Entrada"].isna().all() else grupo.index[0]
                    ultima_data = grupo.loc[ultima_idx, "Dt Entrada"]
                    ultimo_consultor = grupo.loc[ultima_idx, "Consultor Interno"] if "Consultor Interno" in grupo.columns else ""
                else:
                    ultima_data = None
                    ultimo_consultor = ""
                
                # Criar registro para o resultado
                registro = {
                    "Cliente": cliente,
                    "Nome Cliente": cliente_info["Nome Cliente"],
                    "ABC": cliente_info["ABC"],
                    "UF": cliente_info["UF"],
                    "Cidade": cliente_info["Cidade"],
                    "Valor Total Orçado": cliente_info["Valor Total Orçado"],
                    "Código Produto": codigo_produto,
                    "Descrição Produto": produto_info["Descrição Produto"],
                    "Dt Entrada": historico_datas,
                    "Prob.Fech.": historico_prob,
                    "Motivo Não Venda": historico_motivo,
                    "Última Data": ultima_data,
                    "Último Consultor": ultimo_consultor
                }
                
                # Adicionar categorias se disponíveis
                for cat_col in colunas_disponiveis:
                    registro[cat_col] = produto_info[cat_col]
                
                resultado.append(registro)
            
            # Liberar memória
            import gc
            gc.collect()
        
        # Limpar elementos de progresso
        progress_bar.empty()
        status_texto.empty()
        
        # 8. Criar DataFrame final
        df_final = pd.DataFrame(resultado)
        
        # 9. Formatar datas
        if "Última Data" in df_final.columns:
            df_final["Última Data"] = pd.to_datetime(df_final["Última Data"]).dt.strftime("%Y-%m-%d")
        
        st.success(f"Processamento concluído! {len(df_final)} registros gerados.")
        return df_final
        
    except Exception as e:
        st.error(f"Erro no processamento de dados: {e}")
        import traceback
        st.error(traceback.format_exc())
        return pd.DataFrame()


@st.cache_data(ttl=600)
def filtrar_dataframe(df, negocio, grupo, subgrupo, cliente, consultor):
    """Filtra o dataframe com base nos critérios selecionados"""
    df_filtrado = df.copy()
    
    if negocio != 'Todos':
        df_filtrado = df_filtrado[df_filtrado['Negócio'] == negocio]
    if grupo != 'Todos':
        df_filtrado = df_filtrado[df_filtrado['Grupo'] == grupo]
    if subgrupo != 'Todos':
        df_filtrado = df_filtrado[df_filtrado['Subgrupo'] == subgrupo]
    if cliente != 'Todos':
        df_filtrado = df_filtrado[df_filtrado['Nome Cliente'] == cliente]
    if consultor != 'Todos':
        df_filtrado = df_filtrado[df_filtrado['Último Consultor'] == consultor]
    
    return df_filtrado

@st.cache_data(ttl=600)
def ordenar_dataframe(df, coluna, ascendente=True):
    """Ordena o dataframe pela coluna especificada"""
    return df.sort_values(by=coluna, ascending=ascendente)




# Helper functions
def classificar_clientes_abc(df):
    """Classifica os clientes conforme análise ABC baseada no valor orçado."""
    try:
        st.write("Iniciando classificação ABC de clientes...")
        
        # Verificar se temos as colunas necessárias
        if "Cliente" not in df.columns or "Valor Orçado" not in df.columns:
            st.warning("Não é possível realizar classificação ABC: Colunas 'Cliente' ou 'Valor Orçado' ausentes.")
            # Retornar um DataFrame vazio com as colunas necessárias
            return pd.DataFrame(columns=["Cliente", "ABC", "UF", "Cidade", "Valor Total Orçado"])
        
        # Agrupar valores por cliente
        df_clientes = df.groupby("Cliente").agg({
            "Valor Orçado": "sum",
            "Nome Cliente": "first",
            "UF": "first" if "UF" in df.columns else lambda x: "",
            "Cidade": "first" if "Cidade" in df.columns else lambda x: ""
        }).reset_index()
        
        # Renomear colunas
        df_clientes = df_clientes.rename(columns={"Valor Orçado": "Valor Total Orçado"})
        
        # Ordenar por valor total (decrescente)
        df_clientes = df_clientes.sort_values("Valor Total Orçado", ascending=False)
        
        # Calcular valor total e percentuais
        valor_total = df_clientes["Valor Total Orçado"].sum()
        
        if valor_total == 0:
            st.warning("Valor total orçado é zero. Não é possível fazer classificação ABC.")
            df_clientes["ABC"] = "C"
            df_clientes["Percentual"] = 0
            df_clientes["Percentual Acumulado"] = 0
            return df_clientes
        
        df_clientes["Percentual"] = df_clientes["Valor Total Orçado"] / valor_total * 100
        df_clientes["Percentual Acumulado"] = df_clientes["Percentual"].cumsum()
        
        # Classificação ABC
        df_clientes["ABC"] = "C"
        df_clientes.loc[df_clientes["Percentual Acumulado"] <= 80, "ABC"] = "A"
        df_clientes.loc[(df_clientes["Percentual Acumulado"] > 80) & 
                       (df_clientes["Percentual Acumulado"] <= 95), "ABC"] = "B"
        
        # Formatação final
        df_clientes['Ranking'] = df_clientes['Valor Total Orçado'].rank(ascending=False, method='min').astype(int)
        
        st.success(f"Classificação ABC concluída: {len(df_clientes)} clientes classificados")
        
        return df_clientes
    
    except Exception as e:
        st.error(f"Erro ao processar classificação ABC: {e}")
        import traceback
        st.error(traceback.format_exc())
        return pd.DataFrame(columns=["Cliente", "ABC", "UF", "Cidade", "Valor Total Orçado"])
    
    

def juntar_categorias_produtos(df, df_categorias):
    """Realiza junção dos dados de produtos com suas categorias."""
    try:
        df_categorias_slim = df_categorias[["Código Produto", "Negócio", "Grupo", "Subgrupo"]]
        return pd.merge(df, df_categorias_slim, on="Código Produto", how="left")
    except Exception as e:
        st.error(f"Erro ao juntar categorias: {e}")
        return df

@st.cache_data(ttl=600)
def paginar_dataframe(df, page, items_per_page):
    """Retorna apenas os dados da página solicitada"""
    start_idx = (page - 1) * items_per_page
    end_idx = min(start_idx + items_per_page, len(df))
    return df.iloc[start_idx:end_idx].copy()


def diagnosticar_dados(df):
    """Verifica o dataframe em busca de problemas comuns."""
    problemas = []
    
    # Verificar colunas necessárias
    colunas_necessarias = ["Cliente", "Código Produto", "Dt Entrada", "Valor Orçado"]
    for coluna in colunas_necessarias:
        if coluna not in df.columns:
            problemas.append(f"Coluna '{coluna}' não encontrada")
    
    # Verificar valores ausentes em colunas críticas
    for coluna in [c for c in colunas_necessarias if c in df.columns]:
        nulos = df[coluna].isna().sum()
        if nulos > 0:
            problemas.append(f"Coluna '{coluna}' tem {nulos} valores ausentes")
    
    # Verificar tipos de dados
    if "Dt Entrada" in df.columns and not pd.api.types.is_datetime64_dtype(df["Dt Entrada"]):
        problemas.append("Coluna 'Dt Entrada' não está no formato datetime")
    
    if "Valor Orçado" in df.columns:
        try:
            pd.to_numeric(df["Valor Orçado"], errors='raise')
        except:
            problemas.append("Coluna 'Valor Orçado' contém valores que não são numéricos")
    
    # Resumo
    if problemas:
        st.warning("Problemas encontrados nos dados:")
        for problema in problemas:
            st.write(f"- {problema}")
    else:
        st.success("Dados verificados com sucesso!")
    
    # Mostrar informações do dataframe
    st.write(f"Dimensões: {df.shape[0]} linhas x {df.shape[1]} colunas")
    st.write(f"Colunas: {', '.join(map(str, df.columns))}")
    
    return len(problemas) == 0


def limpar_dataframe(df):
    """Limpa e prepara o dataframe para processamento."""
    try:
        linhas_originais = len(df)
        colunas_originais = len(df.columns)
        
        st.write(f"Iniciando limpeza de dados: {linhas_originais} linhas × {colunas_originais} colunas")
        
        # 1. Remover colunas vazias ou com nomes problemáticos
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]  # Remove colunas 'Unnamed'
        
        # 2. Renomear colunas duplicadas para evitar ambiguidades
        # Encontrar colunas com .1, .2, etc. e substituir pelos nomes principais
        rename_cols = {}
        for col in df.columns:
            if re.search(r'\.\d+$', col):
                base_name = re.sub(r'\.\d+$', '', col)
                # Se a coluna base já existe, vamos manter a duplicata com outro nome
                if base_name in df.columns:
                    continue
                else:
                    rename_cols[col] = base_name
        
        # Aplicar renomeação
        df = df.rename(columns=rename_cols)
        
        # 3. Identificar colunas essenciais
        colunas_essenciais = ["Cliente", "Código Produto", "Dt Entrada", "Valor Orçado"]
        
        # 4. Filtrar linhas com dados essenciais (remover linhas onde todas as colunas essenciais são nulas)
        df_limpo = df.dropna(subset=[col for col in colunas_essenciais if col in df.columns], how='all')
        
        # 5. Preencher valores ausentes
        if "Valor Orçado" in df_limpo.columns:
            df_limpo["Valor Orçado"] = df_limpo["Valor Orçado"].fillna(0)
        
        if "Prob.Fech." in df_limpo.columns:
            df_limpo["Prob.Fech."] = df_limpo["Prob.Fech."].fillna(0)
        
        # 6. Converter tipos de dados
        if "Dt Entrada" in df_limpo.columns:
            df_limpo["Dt Entrada"] = pd.to_datetime(df_limpo["Dt Entrada"], errors='coerce')
        
        # 7. Remover linhas duplicadas
        df_limpo = df_limpo.drop_duplicates(subset=[col for col in ["Cliente", "Código Produto", "Dt Entrada"] 
                                                  if col in df_limpo.columns])
        
        # Relatório de limpeza
        linhas_removidas = linhas_originais - len(df_limpo)
        colunas_removidas = colunas_originais - len(df_limpo.columns)
        
        st.success(f"Limpeza concluída: {linhas_removidas} linhas removidas, {colunas_removidas} colunas removidas")
        st.write(f"DataFrame limpo: {len(df_limpo)} linhas × {len(df_limpo.columns)} colunas")
        
        # Mostrar informações sobre o DF limpo
        if len(df_limpo) > 0:
            info = {
                "Total de clientes": df_limpo["Cliente"].nunique() if "Cliente" in df_limpo.columns else 0,
                "Total de produtos": df_limpo["Código Produto"].nunique() if "Código Produto" in df_limpo.columns else 0,
                "Período": f"{df_limpo['Dt Entrada'].min():%Y-%m-%d} a {df_limpo['Dt Entrada'].max():%Y-%m-%d}" 
                        if "Dt Entrada" in df_limpo.columns else "N/A",
                "Total orçado": f"R$ {df_limpo['Valor Orçado'].sum():,.2f}" 
                              if "Valor Orçado" in df_limpo.columns else 0
            }
            
            for k, v in info.items():
                st.write(f"**{k}:** {v}")
        
        return df_limpo
        
    except Exception as e:
        st.error(f"Erro durante a limpeza de dados: {e}")
        import traceback
        st.error(traceback.format_exc())
        return df  # Retorna o DataFrame original em caso de erro


def verificar_estrutura_excel(arquivo):
    """
    Verifica a estrutura de um arquivo Excel para identificar problemas de formatação.
    
    Args:
        arquivo: Caminho do arquivo ou objeto de arquivo
        
    Returns:
        Informações sobre a estrutura do arquivo
    """
    try:
        # Criar um dicionário para armazenar informações
        info = {}
        
        # Se for um arquivo carregado pelo Streamlit, precisamos criar uma cópia temporária
        if hasattr(arquivo, 'read'):
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(arquivo.getvalue())
                tmp_path = tmp.name
                arquivo_path = tmp_path
        else:
            arquivo_path = arquivo
        
        # Usar openpyxl para examinar o arquivo mais detalhadamente
        import openpyxl
        wb = openpyxl.load_workbook(arquivo_path, data_only=True)
        sheet = wb.active
        
        # Obter informações básicas
        info['total_rows'] = sheet.max_row
        info['total_cols'] = sheet.max_column
        
        # Verificar as primeiras linhas para entender a estrutura
        first_rows = []
        for i in range(1, min(6, sheet.max_row + 1)):
            row = []
            for j in range(1, min(11, sheet.max_column + 1)):
                cell_value = sheet.cell(row=i, column=j).value
                row.append(str(cell_value) if cell_value is not None else '')
            first_rows.append(row)
        
        info['first_rows'] = first_rows
        
        # Sugerir qual linha deve ser o cabeçalho
        # Geralmente é a primeira linha, mas podemos tentar detectar automaticamente
        header_candidates = []
        for i in range(0, min(5, len(first_rows))):
            row = first_rows[i]
            # Verificar se parece um cabeçalho (não tem valores numéricos, etc)
            is_header = all(not str(cell).replace('.', '').isdigit() for cell in row if cell)
            if is_header:
                header_candidates.append(i)
        
        info['suggested_header'] = header_candidates[0] if header_candidates else 0
        
        # Limpar arquivos temporários
        if hasattr(arquivo, 'read') and 'tmp_path' in locals():
            import os
            os.unlink(tmp_path)
        
        return info
        
    except Exception as e:
        st.error(f"Erro ao verificar estrutura do arquivo: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return {"error": str(e)}
    
    
# Adicione estas funções auxiliares
def formatar_tupla_dados(tupla_dados):
    """Formata a tupla de dados para exibição legível"""
    if isinstance(tupla_dados, (list, tuple)):
        return '\n\n'.join(map(str, tupla_dados))
    return str(tupla_dados)

def converter_listas_para_visualizacao(df):
    """Converte colunas de listas para formato legível no Streamlit"""
    df_viz = df.copy()
    for col in df.columns:
        if isinstance(df[col].iloc[0], list):
            df_viz[col] = df_viz[col].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
    return df_viz



# Modifique a função verificar_compatibilidade_dataframes para ser menos restritiva:

def verificar_compatibilidade_dataframes(df_analise, df_categorias):
    """Verifica se os dataframes são compatíveis para processamento conjunto."""
    problemas = []
    avisos = []
    
    # Verificar apenas a coluna de ligação (única realmente necessária)
    if "Código Produto" not in df_analise.columns:
        problemas.append("Coluna 'Código Produto' não encontrada no DataFrame de análise")
    
    if "Código Produto" not in df_categorias.columns:
        problemas.append("Coluna 'Código Produto' não encontrada no DataFrame de categorias")
    
    # As outras colunas são esperadas apenas em seus respectivos DataFrames
    colunas_esperadas_categorias = ["Negócio", "Grupo", "Subgrupo"]
    faltantes_categorias = [c for c in colunas_esperadas_categorias if c not in df_categorias.columns]
    
    if faltantes_categorias:
        avisos.append(f"Algumas colunas de categorização não encontradas: {', '.join(faltantes_categorias)}")
    
    # Verificar correspondência entre produtos (como aviso, não problema crítico)
    if "Código Produto" in df_analise.columns and "Código Produto" in df_categorias.columns:
        produtos_analise = set(df_analise["Código Produto"].unique())
        produtos_categorias = set(df_categorias["Código Produto"].unique())
        
        produtos_sem_categoria = produtos_analise - produtos_categorias
        
        if len(produtos_sem_categoria) > 0:
            n_produtos_sem_cat = len(produtos_sem_categoria)
            pct_produtos_sem_cat = n_produtos_sem_cat / len(produtos_analise) * 100
            avisos.append(f"{n_produtos_sem_cat} produtos ({pct_produtos_sem_cat:.1f}%) não têm correspondência no DataFrame de categorias")
    
    # Exibir alertas
    if problemas:
        st.error("Problemas críticos de compatibilidade encontrados:")
        for problema in problemas:
            st.write(f"- {problema}")
        st.error("Estes problemas impedem o processamento conjunto dos dados.")
    
    if avisos:
        st.warning("Avisos sobre compatibilidade:")
        for aviso in avisos:
            st.write(f"- {aviso}")
        st.info("Estes avisos não impedem o processamento, mas a análise pode estar incompleta.")
    
    if not problemas and not avisos:
        st.success("DataFrames são compatíveis para processamento conjunto.")
    
    # Retornar True se não houver problemas críticos
    return len(problemas) == 0




# Sidebar for file upload
st.sidebar.header("Upload de Arquivos")
# Expandir opções avançadas
with st.sidebar.expander("Opções avançadas de carregamento", expanded=True):
    header_analise = st.number_input("Cabeçalho do arquivo de análise (linha):", 0, 10, 0)
    header_categorias = st.number_input("Cabeçalho do arquivo de categorias (linha):", 0, 10, 0)

arquivo_analise = st.sidebar.file_uploader("Arquivo de Análise Comercial", type=["xlsx"])
arquivo_categorias = st.sidebar.file_uploader("Arquivo de Classificação de Produtos", type=["xlsx"])

# Adicione aqui o controle para amostras menores
with st.sidebar.expander("Configurações de Desenvolvimento", expanded=False):
    modo_dev = st.checkbox("Modo desenvolvimento (limitar dados)", value=False)
    if modo_dev:
        max_linhas = st.slider("Máximo de linhas a processar", 100, 10000, 5000)

# Initialize DataFrame variables
df_analise = None
df_categorias = None
df_final = None

# Main app logic
if arquivo_analise is not None and arquivo_categorias is not None:
    # Load data using the proper header rows
    with st.spinner("Carregando arquivos..."):
        # Usar a nova função com os cabeçalhos definidos pelo usuário
        df_analise = carregar_excel_corretamente(arquivo_analise, header_row=header_analise)
        df_categorias = carregar_excel_corretamente(arquivo_categorias, header_row=header_categorias)
        
        if arquivo_analise is not None:
            
            # Mostrar diagnóstico inicial
            with st.expander("Pré-visualização dos dados carregados", expanded=True):
                st.subheader("Primeiras linhas do arquivo de análise")
                st.dataframe(df_analise.head())
                
                st.subheader("Primeiras linhas do arquivo de categorias") 
                st.dataframe(df_categorias.head())
                
                st.subheader("Verificação de compatibilidade")
                verificar_compatibilidade_dataframes(df_analise, df_categorias)
            
            # Limitar dados se estiver em modo desenvolvedor
            if modo_dev:
                df_analise = df_analise.head(max_linhas)
                st.info(f"Modo desenvolvimento: processando apenas {len(df_analise)} linhas")
            
            # Opção para continuar com o processamento
            # Após o processamento dos dados
            if st.button("Processar dados"):
                with st.spinner("Processando dados..."):
                    df_final = processar_dados(df_analise, df_categorias)
                    if df_final is not None and len(df_final) > 0:
                        st.session_state.df_final = df_final
                        st.success(f"Processamento concluído! {len(df_final)} registros disponíveis para análise.")
                        
                        # Não use experimental_rerun() - ele interrompe o fluxo
                        # Em vez disso, defina uma flag para mostrar as tabs na mesma execução
                        st.session_state.mostrar_tabs = True
                    else:
                        st.error("Não foi possível processar os dados corretamente.")

            # Modificar a verificação para exibir as abas
            if ('df_final' in st.session_state and not st.session_state.df_final.empty) or \
            ('mostrar_tabs' in st.session_state and st.session_state.mostrar_tabs):
                
                # Garantir que temos o DataFrame mais recente
                df_final = st.session_state.df_final
                
                # Título e tabs - Colocar em um bloco separado do if
                st.header("Dashboard de Análise")
                
                # Tab selection
                tab_names = ["Visualização de Dados", "Análise Estatística", "Análise Avançada", "Propostas Pendentes"]
                current_tab = st.radio("Selecione a visualização:", tab_names, horizontal=True, 
                                    index=st.session_state.get('current_tab', 0))
                st.session_state.current_tab = tab_names.index(current_tab)
            
            with st.expander("Verificar estrutura do arquivo de análise", expanded=True):
                if st.button("Analisar estrutura do arquivo"):
                    estrutura = verificar_estrutura_excel(arquivo_analise)
                    
                    st.write(f"Total de linhas: {estrutura.get('total_rows', 'N/A')}")
                    st.write(f"Total de colunas: {estrutura.get('total_cols', 'N/A')}")
                    
                    st.subheader("Visualização das primeiras linhas")
                    
                    # Criar uma tabela com as primeiras linhas
                    if 'first_rows' in estrutura:
                        import pandas as pd
                        df_preview = pd.DataFrame(estrutura['first_rows'])
                        st.dataframe(df_preview)
                        
                        # Sugerir o cabeçalho
                        st.info(f"Linha sugerida para cabeçalho: {estrutura.get('suggested_header', 0)}")
                        
                        # Adicionar botão para usar esta sugestão
                        if st.button("Usar linha sugerida como cabeçalho"):
                            header_analise = estrutura.get('suggested_header', 0)
    
    if df_analise is not None and df_categorias is not None:
        # Mostrar diagnóstico inicial
        with st.expander("Pré-visualização dos dados carregados", expanded=True):
            st.subheader("Primeiras linhas do arquivo de análise")
            st.dataframe(df_analise.head())
            
            st.subheader("Primeiras linhas do arquivo de categorias") 
            st.dataframe(df_categorias.head())
            
            # Mostrar as colunas de cada dataframe
            st.subheader("Colunas do arquivo de análise")
            st.write(df_analise.columns.tolist())
            
            st.subheader("Colunas do arquivo de categorias")
            st.write(df_categorias.columns.tolist())
        
        # Limitar dados se estiver em modo desenvolvedor
        if modo_dev:
            df_analise = df_analise.head(max_linhas)
            st.info(f"Modo desenvolvimento: processando apenas {len(df_analise)} linhas")
        
        # Adicionar botão para iniciar o processamento
        with st.expander("Diagnóstico dos dados", expanded=True):
            st.subheader("DataFrame de Análise")
            # Para o DataFrame de análise, verificamos colunas relevantes para análise
            problemas_analise = []
            
            # Verificar colunas necessárias para análise comercial
            colunas_analise = ["Cliente", "Código Produto", "Dt Entrada", "Valor Orçado"]
            for coluna in colunas_analise:
                if coluna not in df_analise.columns:
                    problemas_analise.append(f"Coluna '{coluna}' não encontrada")
            
            if problemas_analise:
                st.warning("Problemas encontrados nos dados:")
                for problema in problemas_analise:
                    st.write(f"- {problema}")
            else:
                st.success("DataFrame de Análise verificado com sucesso!")
            
            st.write(f"Dimensões: {df_analise.shape[0]} linhas x {df_analise.shape[1]} colunas")
            st.write(f"Colunas: {', '.join(map(str, df_analise.columns))}")
            
            st.subheader("DataFrame de Categorias")
            # Para o DataFrame de categorias, verificamos apenas colunas de categorização
            problemas_categorias = []
            
            # A única coluna realmente necessária é a de ligação
            if "Código Produto" not in df_categorias.columns:
                problemas_categorias.append("Coluna 'Código Produto' não encontrada")
            
            # As demais são opcionais para categorização
            colunas_opcionais = ["Negócio", "Grupo", "Subgrupo"]
            colunas_faltantes = [c for c in colunas_opcionais if c not in df_categorias.columns]
            if colunas_faltantes:
                problemas_categorias.append(f"Colunas opcionais não encontradas: {', '.join(colunas_faltantes)}")
            
            if problemas_categorias:
                st.warning("Problemas encontrados nos dados:")
                for problema in problemas_categorias:
                    st.write(f"- {problema}")
            else:
                st.success("DataFrame de Categorias verificado com sucesso!")
            
            st.write(f"Dimensões: {df_categorias.shape[0]} linhas x {df_categorias.shape[1]} colunas")
            st.write(f"Colunas: {', '.join(map(str, df_categorias.columns))}")

# Substitua o bloco de código onde aparecem as abas com este código:

# Tabs system - only show if data has been processed
if df_final is not None or ('df_final' in st.session_state and st.session_state.df_final is not None):
    # Use the stored df_final if available
    if df_final is not None and not df_final.empty:
        st.session_state.df_final = df_final
    
    # Garantir que temos o DataFrame mais recente
    df_final = st.session_state.df_final
    
    # Initialize current tab if not in session state
    if 'current_tab' not in st.session_state:
        st.session_state.current_tab = 0

    # Título da seção principal
    st.header("Dashboard de Análise")

    # Tab selection - Coloque aqui para garantir que as abas apareçam
    tab_names = ["Visualização de Dados", "Análise Estatística", "Análise Avançada", "Propostas Pendentes"]
    current_tab = st.radio("Selecione a aba:", tab_names, horizontal=True, 
                           index=st.session_state.current_tab)
    st.session_state.current_tab = tab_names.index(current_tab)

    # === CONTEÚDO DAS ABAS ===
    
    # === PRIMEIRA ABA: VISUALIZAÇÃO DE DADOS ===
    if current_tab == "Visualização de Dados":
        st.subheader("Análise de Produtos por Cliente")
        
        # Recuperar o DataFrame da sessão
        if 'df_final' in st.session_state and st.session_state.df_final is not None:
            df_final = st.session_state.df_final
            
            # Verificar quais colunas estão disponíveis (sem gerar avisos)
            colunas_disponiveis = df_final.columns.tolist()
            colunas_esperadas = [
                "Negócio", "Grupo", "Subgrupo", 
                "Cliente", "Nome Cliente", "ABC", "UF", "Cidade",
                "Código Produto", "Descrição Produto", 
                "Última Data", "Último Consultor", "Valor Total Orçado"
            ]
            
            # Silenciosamente adicionar as colunas faltantes
            for coluna in colunas_esperadas:
                if coluna not in colunas_disponiveis:
                    df_final[coluna] = ""
            
        
        # Opção para verificar o DataFrame (agora dentro da primeira aba)
        with st.expander("📊 Verificar DataFrame Final", expanded=False):
            st.subheader("Informações do DataFrame Final")
            
            # 1. Informações básicas
            st.write(f"**Dimensões:** {df_final.shape[0]} linhas × {df_final.shape[1]} colunas")
            
            # 2. Lista de colunas existentes
            st.write("**Colunas disponíveis:**")
            colunas_disponiveis = df_final.columns.tolist()
            st.write(", ".join(colunas_disponiveis))
            
            # 3. Verificar colunas essenciais
            colunas_essenciais = [
                "Negócio", "Grupo", "Subgrupo", 
                "Cliente", "Nome Cliente", "ABC", "UF", "Cidade",
                "Código Produto", "Descrição Produto", 
                "Última Data", "Último Consultor", "Valor Total Orçado",
                "Dt Entrada", "Prob.Fech.", "Motivo Não Venda"
            ]
            
            colunas_faltantes = [col for col in colunas_essenciais if col not in colunas_disponiveis]
            
            if colunas_faltantes:
                st.warning(f"**Colunas essenciais faltando:** {', '.join(colunas_faltantes)}")
            else:
                st.success("Todas as colunas essenciais estão presentes!")
            
            # 4. Visualizar as primeiras linhas
            st.write("**Primeiras 5 linhas:**")
            st.dataframe(df_final.head())
            
            # 5. Verificar tipos de dados
            st.write("**Tipos de dados:**")
            tipos = df_final.dtypes.reset_index()
            tipos.columns = ["Coluna", "Tipo"]
            st.dataframe(tipos)
            
            # 6. Permitir correção manual
            if st.checkbox("Precisa corrigir manualmente o DataFrame?"):
                st.warning("Ajustes manuais podem ser necessários se o processamento não incluiu todas as colunas necessárias.")
                
                # Opção para adicionar coluna faltante
                if colunas_faltantes:
                    col_to_add = st.selectbox("Selecione uma coluna para adicionar:", colunas_faltantes)
                    
                    if st.button(f"Adicionar coluna {col_to_add} com valores vazios"):
                        df_final[col_to_add] = ""
                        st.session_state.df_final = df_final
                        st.success(f"Coluna {col_to_add} adicionada! Recarregue a página para ver as mudanças.")
                
                # Opção para executar código personalizado
                st.write("**Executar código personalizado para corrigir o DataFrame:**")
                custom_code = st.text_area("Digite o código Python (use df_final como nome do DataFrame):", 
                                        "# Exemplo:\n# df_final['Nova Coluna'] = df_final['Coluna Existente']\n# df_final.rename(columns={'Nome Antigo': 'Nome Novo'}, inplace=True)", 
                                        height=150)
                
                if st.button("Executar código"):
                    try:
                        # Criar cópia segura para evitar modificações indesejadas
                        df_temp = df_final.copy()
                        
                        # Executar o código inserido
                        exec(custom_code, {"df_final": df_temp, "pd": pd, "np": np})
                        
                        # Verificar se o código foi executado sem erros
                        st.success("Código executado com sucesso!")
                        
                        # Mostrar o resultado
                        st.write("**Resultado após execução do código:**")
                        st.dataframe(df_temp.head())
                        
                        # Opção para salvar as alterações
                        if st.button("Salvar alterações ao DataFrame"):
                            df_final = df_temp
                            st.session_state.df_final = df_final
                            st.success("Alterações salvas! Recarregue a página para ver as mudanças.")
                            
                    except Exception as e:
                        st.error(f"Erro ao executar código: {str(e)}")
        
        # Verificar e adicionar colunas faltantes
        colunas_essenciais = [
            "Negócio", "Grupo", "Subgrupo", 
            "Cliente", "Nome Cliente", "ABC", "UF", "Cidade",
            "Código Produto", "Descrição Produto", 
            "Última Data", "Último Consultor", "Valor Total Orçado",
            "Dt Entrada", "Prob.Fech.", "Motivo Não Venda"
        ]
        
        for coluna in colunas_essenciais:
            if coluna not in df_final.columns:
                df_final[coluna] = ""
                st.warning(f"Coluna '{coluna}' não encontrada nos dados. Adicionada com valores vazios.")
        
        # Métricas resumidas
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total de Registros", len(df_final))
        with col2:
            st.metric("Total de Clientes", df_final["Cliente"].nunique())
        with col3:
            st.metric("Total de Produtos", df_final["Código Produto"].nunique())
        with col4:
            st.metric("Clientes A", len(df_final[df_final["ABC"] == "A"]["Cliente"].unique()))
        
        # Resto do conteúdo da primeira aba...
        # [Mantenha o código existente para filtros, etc.]

    # === SEGUNDA ABA: ANÁLISE ESTATÍSTICA ===
    elif current_tab == "Análise Estatística":
        st.header("Análise Estatística")
        # ...resto do código da tab2...

    # === TERCEIRA ABA: ANÁLISE AVANÇADA ===
    elif current_tab == "Análise Avançada":
        st.header("Análise Avançada")
        # ...resto do código da tab3...

    # === QUARTA ABA: PROPOSTAS PENDENTES ===
    elif current_tab == "Propostas Pendentes":
        # Chame a função do módulo importado
        exibir_analise_pendentes()

else:
    st.info("Por favor, faça o upload dos arquivos de dados para visualizar a análise comercial.")
